import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from tkinter.scrolledtext import ScrolledText
import threading
import sys
import io
from contextlib import redirect_stdout, redirect_stderr
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
import re
import os
import webbrowser


# Utility to get resource path for PyInstaller and normal script
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class AuditReportGUI:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.setup_variables()
        self.create_menu()
        self.create_widgets()
        
    def setup_window(self):
        self.root.title("TNBT Post Drainage Final Report Generator - Developed by Rishav Raj")
        self.root.geometry("900x700")
        self.root.configure(bg='#1e1e2e')
        self.root.resizable(True, True)
        
        # Configure style for modern look
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors for ttk widgets
        style.configure('Company.TLabel', 
                       background='#1e1e2e', 
                       foreground='#f9e2af', 
                       font=('Segoe UI', 20, 'bold'))
        style.configure('Title.TLabel', 
                       background='#1e1e2e', 
                       foreground='#cdd6f4', 
                       font=('Segoe UI', 16, 'bold'))
        style.configure('Subtitle.TLabel', 
                       background='#1e1e2e', 
                       foreground='#a6adc8', 
                       font=('Segoe UI', 10))
        style.configure('Footer.TLabel',
                       background='#1e1e2e',
                       foreground='#a6adc8',
                       font=('Segoe UI', 9, 'italic'))
        style.configure('Modern.TButton',
                       background='#89b4fa',
                       foreground='#1e1e2e',
                       font=('Segoe UI', 10, 'bold'),
                       borderwidth=0,
                       focuscolor='none')
        style.map('Modern.TButton',
                 background=[('active', '#74c0fc'),
                            ('pressed', '#5c7cfa')])
    
    def create_menu(self):
        """Create the menu bar with developer info and help"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Select Countsheet File", command=lambda: self.browse_file(self.countsheet_path), accelerator="Ctrl+C")
        file_menu.add_command(label="Select Master File", command=lambda: self.browse_file(self.master_path), accelerator="Ctrl+M")
        file_menu.add_command(label="Set Output Location", command=self.browse_save_file, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit, accelerator="Ctrl+Q")
        
        # Tools Menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="🚀 Generate Report", command=self.start_generation, accelerator="F5")
        tools_menu.add_command(label="🗑️ Clear All Fields", command=self.clear_all, accelerator="Ctrl+L")
        tools_menu.add_command(label="📋 Clear Console", command=self.clear_console, accelerator="Ctrl+Shift+L")
        
        # Help Menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="📖 How to Use", command=self.show_help)
        help_menu.add_command(label="📋 About Project", command=self.show_about_project)
        help_menu.add_separator()
        help_menu.add_command(label="👨‍💻 Developer Info", command=self.show_developer_info)
        help_menu.add_command(label="📞 Contact Developer", command=self.show_contact_info)
        help_menu.add_separator()
        help_menu.add_command(label="ℹ️ About", command=self.show_about)
        
        # Bind keyboard shortcuts
        self.root.bind('<Control-c>', lambda e: self.browse_file(self.countsheet_path))
        self.root.bind('<Control-m>', lambda e: self.browse_file(self.master_path))
        self.root.bind('<Control-s>', lambda e: self.browse_save_file())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<F5>', lambda e: self.start_generation())
        self.root.bind('<Control-l>', lambda e: self.clear_all())
        self.root.bind('<Control-Shift-L>', lambda e: self.clear_console())
        
    def setup_variables(self):
        self.countsheet_path = tk.StringVar()
        self.template_path = resource_path('template.xlsx')  # Use resource_path for template
        self.master_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.processing = False
        
    def create_widgets(self):
        # Main container
        main_frame = tk.Frame(self.root, bg='#1e1e2e')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Company Name at Top Center
        company_label = ttk.Label(main_frame, 
                                 text="TNBT", 
                                 style='Company.TLabel',
                                 anchor='center',
                                 justify='center')
        company_label.pack(pady=(0, 2))
        
        # Title
        title_label = ttk.Label(main_frame, 
                               text="🔍 Post Drainage Final Report Generator", 
                               style='Title.TLabel',
                               anchor='center',
                               justify='center')
        title_label.pack(pady=(0, 5))
        
        subtitle_label = ttk.Label(main_frame, 
                                  text="Generate comprehensive audit reports from Excel data", 
                                  style='Subtitle.TLabel')
        subtitle_label.pack(pady=(0, 20))
        
        # File selection frame
        files_frame = tk.Frame(main_frame, bg='#1e1e2e')
        files_frame.pack(fill='x', pady=(0, 20))
        
        # Input files section
        self.create_file_section(files_frame, "📊 Input Files", [
            ("Countsheet File:", self.countsheet_path, "Select the countsheet Excel file"),
            ("Master File:", self.master_path, "Select the master lookup Excel file")
        ])
        
        # Output section
        self.create_output_section(files_frame)
        
        # Control buttons
        self.create_control_buttons(main_frame)
        
        # Console section
        self.create_console_section(main_frame)
        
        # Developer credit footer
        self.create_footer()
    
    def create_file_section(self, parent, title, file_configs):
        section_frame = tk.LabelFrame(parent, 
                                    text=title, 
                                    bg='#313244', 
                                    fg='#cdd6f4',
                                    font=('Segoe UI', 11, 'bold'),
                                    relief='flat',
                                    bd=2)
        section_frame.pack(fill='x', pady=(0, 15))
        
        for label_text, var, tooltip in file_configs:
            self.create_file_input(section_frame, label_text, var, tooltip)
    
    def create_file_input(self, parent, label_text, var, tooltip):
        frame = tk.Frame(parent, bg='#313244')
        frame.pack(fill='x', padx=15, pady=8)
        
        # Label
        label = tk.Label(frame, 
                        text=label_text, 
                        bg='#313244', 
                        fg='#cdd6f4',
                        font=('Segoe UI', 10),
                        width=15,
                        anchor='w')
        label.pack(side='left')
        
        # Entry
        entry = tk.Entry(frame, 
                        textvariable=var,
                        bg='#45475a',
                        fg='#cdd6f4',
                        font=('Segoe UI', 10),
                        relief='flat',
                        bd=5)
        entry.pack(side='left', fill='x', expand=True, padx=(10, 5))
        
        # Browse button
        browse_btn = tk.Button(frame,
                              text="Browse",
                              command=lambda: self.browse_file(var),
                              bg='#74c0fc',
                              fg='#1e1e2e',
                              font=('Segoe UI', 9, 'bold'),
                              relief='flat',
                              bd=0,
                              padx=15,
                              cursor='hand2')
        browse_btn.pack(side='right')
        
        # Add hover effects
        def on_enter(e):
            browse_btn.config(bg='#89b4fa')
        def on_leave(e):
            browse_btn.config(bg='#74c0fc')
            
        browse_btn.bind('<Enter>', on_enter)
        browse_btn.bind('<Leave>', on_leave)
    
    def create_output_section(self, parent):
        output_frame = tk.LabelFrame(parent, 
                                   text="💾 Output Settings", 
                                   bg='#313244', 
                                   fg='#cdd6f4',
                                   font=('Segoe UI', 11, 'bold'),
                                   relief='flat',
                                   bd=2)
        output_frame.pack(fill='x', pady=(0, 15))
        
        frame = tk.Frame(output_frame, bg='#313244')
        frame.pack(fill='x', padx=15, pady=8)
        
        label = tk.Label(frame, 
                        text="Output File:", 
                        bg='#313244', 
                        fg='#cdd6f4',
                        font=('Segoe UI', 10),
                        width=15,
                        anchor='w')
        label.pack(side='left')
        
        entry = tk.Entry(frame, 
                        textvariable=self.output_path,
                        bg='#45475a',
                        fg='#cdd6f4',
                        font=('Segoe UI', 10),
                        relief='flat',
                        bd=5)
        entry.pack(side='left', fill='x', expand=True, padx=(10, 5))
        
        browse_btn = tk.Button(frame,
                              text="Save As",
                              command=self.browse_save_file,
                              bg='#a6e3a1',
                              fg='#1e1e2e',
                              font=('Segoe UI', 9, 'bold'),
                              relief='flat',
                              bd=0,
                              padx=15,
                              cursor='hand2')
        browse_btn.pack(side='right')
        
        # Add hover effects
        def on_enter(e):
            browse_btn.config(bg='#94e2d5')
        def on_leave(e):
            browse_btn.config(bg='#a6e3a1')
            
        browse_btn.bind('<Enter>', on_enter)
        browse_btn.bind('<Leave>', on_leave)
    
    def create_control_buttons(self, parent):
        button_frame = tk.Frame(parent, bg='#1e1e2e')
        button_frame.pack(fill='x', pady=(0, 20))
        
        # Generate button
        self.generate_btn = tk.Button(button_frame,
                                     text="🚀 Generate Report",
                                     command=self.start_generation,
                                     bg='#89b4fa',
                                     fg='#1e1e2e',
                                     font=('Segoe UI', 12, 'bold'),
                                     relief='flat',
                                     bd=0,
                                     padx=30,
                                     pady=12,
                                     cursor='hand2')
        self.generate_btn.pack(side='left', padx=(0, 10))
        
        # Clear button
        clear_btn = tk.Button(button_frame,
                             text="🗑️ Clear All",
                             command=self.clear_all,
                             bg='#f38ba8',
                             fg='#1e1e2e',
                             font=('Segoe UI', 12, 'bold'),
                             relief='flat',
                             bd=0,
                             padx=30,
                             pady=12,
                             cursor='hand2')
        clear_btn.pack(side='left', padx=(0, 10))
        
        # Help button
        help_btn = tk.Button(button_frame,
                            text="❓ Help",
                            command=self.show_help,
                            bg='#cba6f7',
                            fg='#1e1e2e',
                            font=('Segoe UI', 12, 'bold'),
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=12,
                            cursor='hand2')
        help_btn.pack(side='left')
        
        # Progress bar
        self.progress = ttk.Progressbar(button_frame, 
                                       mode='indeterminate',
                                       style='TProgressbar')
        self.progress.pack(side='right', fill='x', expand=True, padx=(20, 0))
        
        # Add hover effects
        def generate_hover_enter(e):
            if not self.processing:
                self.generate_btn.config(bg='#74c0fc')
        def generate_hover_leave(e):
            if not self.processing:
                self.generate_btn.config(bg='#89b4fa')
        
        def clear_hover_enter(e):
            clear_btn.config(bg='#eba0ac')
        def clear_hover_leave(e):
            clear_btn.config(bg='#f38ba8')
        
        def help_hover_enter(e):
            help_btn.config(bg='#b4befe')
        def help_hover_leave(e):
            help_btn.config(bg='#cba6f7')
            
        self.generate_btn.bind('<Enter>', generate_hover_enter)
        self.generate_btn.bind('<Leave>', generate_hover_leave)
        clear_btn.bind('<Enter>', clear_hover_enter)
        clear_btn.bind('<Leave>', clear_hover_leave)
        help_btn.bind('<Enter>', help_hover_enter)
        help_btn.bind('<Leave>', help_hover_leave)
    
    def create_console_section(self, parent):
        console_frame = tk.LabelFrame(parent, 
                                    text="📝 Console Output", 
                                    bg='#313244', 
                                    fg='#cdd6f4',
                                    font=('Segoe UI', 11, 'bold'),
                                    relief='flat',
                                    bd=2)
        console_frame.pack(fill='both', expand=True)
        
        # Console text widget
        self.console = ScrolledText(console_frame,
                                   bg='#181825',
                                   fg='#cdd6f4',
                                   font=('Consolas', 10),
                                   relief='flat',
                                   bd=0,
                                   wrap='word',
                                   height=15)
        self.console.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Initial message
        self.log("🎯 TNBT Post Drainage Final Report Generator Ready!")
        self.log("💻 Developed by Rishav Raj - Professional Audit Report Solution")
        self.log("📁 Please select your input files and output location to begin.")
        self.log("=" * 60)
    
    def create_footer(self):
        """Create footer with developer credit"""
        # Add footer at the bottom
        self.root.update_idletasks()
        footer = ttk.Label(self.root, text="💻 Developed by Rishav Raj", 
                          style='Footer.TLabel', anchor='center', justify='center')
        footer.place(relx=0.5, rely=1.0, anchor='s', y=-5)
    
    def show_help(self):
        """Show how to use dialog"""
        help_text = """
📖 HOW TO USE TNBT POST DRAINAGE FINAL REPORT GENERATOR

🔧 SETUP:
1. Click 'Browse' next to Countsheet File to select the main data file
2. Click 'Browse' next to Master File to select the lookup reference file
3. Click 'Save As' to specify where to save the generated report
4. Ensure template.xlsx is in the same directory as the application

⚡ PROCESSING:
1. Click '🚀 Generate Report' button or press F5
2. Monitor progress in the console below
3. Wait for completion message and success notification

📋 FILE REQUIREMENTS:

📊 COUNTSHEET FILE:
• Excel file (.xlsx/.xlsm/.xls) with main audit data
• Must contain columns: Item/SKU Code, Item Name, Distributor code, etc.
• Original QTY, Original Damage, Original Expired columns for quantities
• Item Rate, Manu Date, Expiry Date for calculations

📁 MASTER FILE:
• Excel file with lookup data for distributors and regions
• Must contain: Anchor Code, Distributor Name, Region, Anchor Name
• Audit Std Serial No for report identification
• Additional fields: DB Name, Distributor City, Reported Value

💾 TEMPLATE FILE:
• template.xlsx must be present in application directory
• Contains the report format and formulas
• Has 'Artical level format' and 'Sign Format.' sheets

🔧 FEATURES:
• Automatic data mapping and transformation
• INR calculations based on quantities and rates
• Subtotal calculations and report formatting
• Master file lookup for distributor details
• Date formatting and audit trail generation

⌨️ KEYBOARD SHORTCUTS:
• Ctrl+C: Select Countsheet File
• Ctrl+M: Select Master File
• Ctrl+S: Set Output Location
• F5: Generate Report
• Ctrl+L: Clear All Fields
• Ctrl+Shift+L: Clear Console
• Ctrl+Q: Exit Application

💡 TIPS:
• Ensure all required columns are present in input files
• Check console for detailed processing information
• Verify template.xlsx is not corrupted or modified
• Use 'Clear All' to reset fields for new processing
• Contact developer for support if needed

🔍 TROUBLESHOOTING:
• If "template.xlsx not found" - ensure file is in app directory
• If "no matching row" - verify Anchor Code and Distributor Name match
• If processing fails - check input file formats and required columns
        """
        
        # Create help window
        help_window = tk.Toplevel(self.root)
        help_window.title("How to Use - TNBT Audit Report Generator")
        help_window.geometry("700x600")
        help_window.configure(bg='#1e1e2e')
        help_window.transient(self.root)
        help_window.grab_set()
        
        # Help text area
        help_text_area = scrolledtext.ScrolledText(help_window, wrap=tk.WORD, 
                                                  font=('Segoe UI', 10),
                                                  bg='#181825', fg='#cdd6f4')
        help_text_area.pack(fill='both', expand=True, padx=20, pady=20)
        help_text_area.insert('1.0', help_text)
        help_text_area.configure(state='disabled')
    
    def show_about_project(self):
        """Show about project dialog"""
        about_text = """
📋 ABOUT TNBT POST DRAINAGE FINAL REPORT GENERATOR

🎯 PROJECT OVERVIEW:
This application is designed to automate the generation of post-drainage audit reports for TNBT organization. It streamlines the process of creating comprehensive audit documentation by processing countsheet data and combining it with master file information to produce formatted Excel reports.

🔧 TECHNICAL FEATURES:
• Excel file processing and data transformation
• Automatic data mapping between multiple sources
• INR calculations based on quantities and rates
• Master file lookup and validation
• Template-based report generation
• Subtotal calculations and formatting
• Date handling and audit trail creation
• Real-time progress tracking and logging

💼 BUSINESS VALUE:
• Reduces audit report generation time by 95%
• Eliminates manual calculation errors
• Ensures consistent report formatting
• Provides comprehensive audit documentation
• Supports regulatory compliance requirements
• Enables efficient post-drainage analysis

🏗️ ARCHITECTURE:
• Built with Python and Tkinter for cross-platform compatibility
• Uses OpenPyXL for advanced Excel manipulation
• Multi-threaded processing for responsive UI
• Template-based approach for flexible reporting
• Robust error handling and validation

📊 PROCESSING WORKFLOW:
1. Load and validate countsheet data
2. Filter valid audit entries (non-zero quantities)
3. Apply data mapping to template format
4. Calculate INR values and subtotals
5. Lookup master file for distributor details
6. Update sign format sheet with totals
7. Generate final formatted report

🎯 TARGET USERS:
• Audit teams and inspectors
• Post-drainage analysis specialists
• Financial reporting staff
• Compliance officers
• Management reporting teams

📈 KEY BENEFITS:
• Automated report generation
• Data accuracy and consistency
• Time-efficient processing
• Professional report formatting
• Comprehensive audit trail
• Easy-to-use interface

🔧 TECHNICAL SPECIFICATIONS:
• Python 3.7+ required
• OpenPyXL library for Excel processing
• Tkinter for GUI interface
• Threading for background processing
• Template-based report structure

📈 VERSION HISTORY:
• v1.0: Initial release with core functionality
• Features: Countsheet processing, master lookup, report generation
        """
        
        # Create about window
        about_window = tk.Toplevel(self.root)
        about_window.title("About Project - TNBT Audit Report Generator")
        about_window.geometry("700x600")
        about_window.configure(bg='#1e1e2e')
        about_window.transient(self.root)
        about_window.grab_set()
        
        # About text area
        about_text_area = scrolledtext.ScrolledText(about_window, wrap=tk.WORD,
                                                   font=('Segoe UI', 10),
                                                   bg='#181825', fg='#cdd6f4')
        about_text_area.pack(fill='both', expand=True, padx=20, pady=20)
        about_text_area.insert('1.0', about_text)
        about_text_area.configure(state='disabled')
    
    def show_developer_info(self):
        """Show developer information dialog"""
        dev_info = """
    👨‍💻 DEVELOPER INFORMATION

    🧑‍💼 NAME: Rishav Raj
    🎯 ROLE: Software Developer & Business Automation Specialist

    💼 EXPERTISE:
    • Python Development & Automation
    • Excel Data Processing & Manipulation
    • GUI Application Development
    • Database Design & Optimization
    • Business Process Automation
    • Web Development
    • Web scraping

    🏆 SPECIALIZATIONS:
    • Enterprise Audit Solutions
    • Post-Drainage Analysis Systems
    • Custom Reporting Applications
    • Data Transformation & Validation
    • Template-Based Report Generation

    🔧 TECHNICAL SKILLS:
    • Languages: Python, SQL, JavaScript, VBA
    • Libraries: OpenPyXL, Pandas, Tkinter, NumPy
    • Tools: Excel, Git, PyInstaller, VS Code
    • Databases: MySQL, SQLite, PostgreSQL
    • Frameworks: Flask, Django, Tkinter

    📈 PROJECT CONTRIBUTIONS:
    • TNBT Post Drainage Report Generator (Current)
    • TNBT Excel Data Processor
    • Multiple audit automation solutions
    • Financial reporting systems
    • Data migration and validation tools

    🎓 EDUCATION & CERTIFICATIONS:
    • Computer Science & Engineering
    • Business Process Automation Specialist
    • Excel Advanced Analytics Certification
    • Python Development Expert

    💡 PHILOSOPHY:
    "Creating intelligent automation solutions that transform complex business processes into simple, efficient workflows while maintaining accuracy and compliance."

    🌟 ACHIEVEMENTS:
    • 1+ years in software development
    • 10+ successful automation projects
    • Expert in Excel data processing
    • Proven track record in enterprise solutions

    🎯 PROJECT FOCUS:
    This Post Drainage Report Generator represents advanced expertise in:
    • Complex Excel data manipulation
    • Multi-source data integration
    • Template-based report automation
    • Audit compliance and documentation
    • User-friendly interface design
        """
        
        # Create developer info window with larger size
        dev_window = tk.Toplevel(self.root)
        dev_window.title("Developer Information - Rishav Raj")
        dev_window.geometry("750x700")  # Increased size
        dev_window.configure(bg='#1e1e2e')
        dev_window.transient(self.root)
        dev_window.grab_set()
        dev_window.resizable(True, True)
        
        # Developer info text area with better sizing
        dev_text_area = scrolledtext.ScrolledText(dev_window, wrap=tk.WORD,
                                                font=('Segoe UI', 10),
                                                bg='#181825', fg='#cdd6f4')
        dev_text_area.pack(fill='both', expand=True, padx=25, pady=(25, 15))
        dev_text_area.insert('1.0', dev_info)
        dev_text_area.configure(state='disabled')
        
        # Contact buttons frame
        contact_frame = tk.Frame(dev_window, bg='#1e1e2e')
        contact_frame.pack(fill='x', padx=25, pady=(0, 25))
        
        contact_btn = tk.Button(contact_frame, text="📞 Contact Info", 
                            command=self.show_contact_info,
                            bg='#89b4fa', fg='#1e1e2e',
                            font=('Segoe UI', 11, 'bold'),
                            relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        contact_btn.pack(side='left', padx=(0, 10))
        
        close_btn = tk.Button(contact_frame, text="❌ Close", 
                            command=dev_window.destroy,
                            bg='#f38ba8', fg='#1e1e2e',
                            font=('Segoe UI', 11, 'bold'),
                            relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        close_btn.pack(side='right')
        
        # Center the window
        dev_window.update_idletasks()
        x = (dev_window.winfo_screenwidth() // 2) - (dev_window.winfo_width() // 2)
        y = (dev_window.winfo_screenheight() // 2) - (dev_window.winfo_height() // 2)
        dev_window.geometry(f"+{x}+{y}")

    
    def show_contact_info(self):
        """Show contact information dialog"""
        # Create contact window with larger size
        contact_window = tk.Toplevel(self.root)
        contact_window.title("Contact Developer - Rishav Raj")
        contact_window.geometry("650x550")  # Increased size
        contact_window.configure(bg='#1e1e2e')
        contact_window.transient(self.root)
        contact_window.grab_set()
        contact_window.resizable(True, True)  # Make it resizable

        # Main frame for header and buttons
        main_frame = tk.Frame(contact_window, bg='#1e1e2e')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Header with larger font
        header_label = tk.Label(main_frame, text="📞 CONTACT DEVELOPER", 
                            font=('Segoe UI', 18, 'bold'),
                            fg='#f9e2af', bg='#1e1e2e')
        header_label.pack(pady=(0, 25))

        # --- Scrollable details area ---
        details_outer_frame = tk.Frame(main_frame, bg='#1e1e2e')
        details_outer_frame.pack(fill='both', expand=True, pady=(0, 25))

        canvas = tk.Canvas(details_outer_frame, bg='#313244', highlightthickness=0)
        scrollbar = tk.Scrollbar(details_outer_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#313244')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Contact items with your updated information
        contact_items = [
            ("👤 Name:", "Rishav Raj"),
            ("📧 Email:", "rishavraj.dev@gmail.com"),
            ("💼 LinkedIn:", "https://www.linkedin.com/in/rishavraj1998/"),
            ("🐙 GitHub:", "https://github.com/rishavraj543256"),
            ("🌐 Portfolio:", "https://rishavraj543256.netlify.app/"),
            ("📱 Phone:", "+91-7903312858"),
            ("🏢 Company:", "TNBT Group"),
            ("📍 Location:", "India")
        ]

        # Create contact items with better spacing
        for i, (label, value) in enumerate(contact_items):
            item_frame = tk.Frame(scrollable_frame, bg='#313244')
            item_frame.pack(fill='x', pady=12)
            label_widget = tk.Label(item_frame, text=label, 
                                font=('Segoe UI', 11, 'bold'),
                                fg='#cdd6f4', bg='#313244',
                                width=12, anchor='w')
            label_widget.pack(side='left')
            if len(value) > 50:
                value_label = tk.Label(item_frame, text=value, 
                                    font=('Segoe UI', 10),
                                    fg='#a6adc8', bg='#313244',
                                    anchor='w', justify='left',
                                    wraplength=400)
            else:
                value_label = tk.Label(item_frame, text=value, 
                                    font=('Segoe UI', 11),
                                    fg='#a6adc8', bg='#313244',
                                    anchor='w')
            value_label.pack(side='left', padx=(15, 0), fill='x', expand=True)
            if "email" in label.lower() or "linkedin" in label.lower() or "github" in label.lower() or "portfolio" in label.lower():
                value_label.configure(fg='#89b4fa', cursor='hand2')
                def on_enter(e, widget=value_label):
                    widget.configure(fg='#74c0fc')
                def on_leave(e, widget=value_label):
                    widget.configure(fg='#89b4fa')
                value_label.bind('<Enter>', on_enter)
                value_label.bind('<Leave>', on_leave)
                if "email" in label.lower():
                    value_label.bind('<Button-1>', lambda e: webbrowser.open(f'mailto:{value}?subject=TNBT Application Support'))
                else:
                    value_label.bind('<Button-1>', lambda e, url=value: webbrowser.open(url))

        # Add a separator line
        separator = tk.Frame(main_frame, height=2, bg='#45475a')
        separator.pack(fill='x', pady=(0, 20))

        # Buttons frame with better styling
        button_frame = tk.Frame(main_frame, bg='#1e1e2e')
        button_frame.pack(fill='x')

        email_btn = tk.Button(button_frame, text="📧 Send Email", 
                            command=lambda: webbrowser.open('mailto:rishavraj.dev@gmail.com?subject=TNBT Application Support&body=Hello Rishav,\n\nI need assistance with the TNBT application.\n\nRegards,'),
                            bg='#89b4fa', fg='#1e1e2e',
                            font=('Segoe UI', 11, 'bold'),
                            relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        email_btn.pack(side='left', padx=(0, 10))
        portfolio_btn = tk.Button(button_frame, text="🌐 Visit Portfolio", 
                                command=lambda: webbrowser.open('https://rishavraj543256.netlify.app/'),
                                bg='#a6e3a1', fg='#1e1e2e',
                                font=('Segoe UI', 11, 'bold'),
                                relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        portfolio_btn.pack(side='left', padx=(0, 10))
        linkedin_btn = tk.Button(button_frame, text="💼 LinkedIn", 
                                command=lambda: webbrowser.open('https://www.linkedin.com/in/rishavraj1998/'),
                                bg='#cba6f7', fg='#1e1e2e',
                                font=('Segoe UI', 11, 'bold'),
                                relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        linkedin_btn.pack(side='left')
        close_btn = tk.Button(button_frame, text="❌ Close", 
                            command=contact_window.destroy,
                            bg='#f38ba8', fg='#1e1e2e',
                            font=('Segoe UI', 11, 'bold'),
                            relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        close_btn.pack(side='right')
        def add_button_hover(button, normal_color, hover_color):
            def on_enter(e):
                button.config(bg=hover_color)
            def on_leave(e):
                button.config(bg=normal_color)
            button.bind('<Enter>', on_enter)
            button.bind('<Leave>', on_leave)
        add_button_hover(email_btn, '#89b4fa', '#74c0fc')
        add_button_hover(portfolio_btn, '#a6e3a1', '#94e2d5')
        add_button_hover(linkedin_btn, '#cba6f7', '#b4befe')
        add_button_hover(close_btn, '#f38ba8', '#eba0ac')
        contact_window.update_idletasks()
        x = (contact_window.winfo_screenwidth() // 2) - (contact_window.winfo_width() // 2)
        y = (contact_window.winfo_screenheight() // 2) - (contact_window.winfo_height() // 2)
        contact_window.geometry(f"+{x}+{y}")
    
    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About TNBT Audit Report Generator", 
                           "TNBT Post Drainage Final Report Generator v1.0\n\n"
                           "Professional audit report automation solution\n"
                           "for post-drainage analysis and compliance.\n\n"
                           "💻 Developed by: Rishav Raj\n"
                           "🏢 Organization: TNBT\n"
                           "📅 Year: 2025\n\n"
                           "Built with Python, Tkinter & OpenPyXL\n"
                           "© 2025 All rights reserved.")
    
    def browse_file(self, var):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            var.set(file_path)
            self.log(f"📄 Selected: {file_path.split('/')[-1]}")
    
    def browse_save_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Report As",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialfile="audit_report.xlsx"
        )
        if file_path:
            self.output_path.set(file_path)
            self.log(f"💾 Output location: {file_path.split('/')[-1]}")
    
    def clear_all(self):
        self.countsheet_path.set("")
        self.template_path = resource_path('template.xlsx') # Use resource_path for template
        self.master_path.set("")
        self.output_path.set("")
        self.console.delete(1.0, tk.END)
        self.log("🧹 All fields cleared!")
        self.log("📁 Please select your input files and output location to begin.")
        self.log("=" * 60)
    
    def clear_console(self):
        """Clear the console"""
        self.console.delete(1.0, tk.END)
        self.log("Console cleared.")
    
    def log(self, message):
        """Add message to console with timestamp"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}\n"
        
        self.console.insert(tk.END, formatted_message)
        self.console.see(tk.END)
        self.root.update_idletasks()
    
    def validate_inputs(self):
        """Validate that all required inputs are provided"""
        if not self.countsheet_path.get():
            return "Please select a Countsheet file"
        if not self.master_path.get():
            return "Please select a Master file"
        if not self.output_path.get():
            return "Please specify output file location"
        return None
    
    def start_generation(self):
        """Start the report generation in a separate thread"""
        error = self.validate_inputs()
        if error:
            messagebox.showerror("Input Error", error)
            return
        
        if self.processing:
            return
        
        self.processing = True
        self.generate_btn.config(text="⏳ Processing...", state='disabled', bg='#6c7086')
        self.progress.start(10)
        
        # Start processing in separate thread
        thread = threading.Thread(target=self.generate_report)
        thread.daemon = True
        thread.start()
    
    def generate_report(self):
        """Main report generation logic"""
        try:
            self.log("🚀 Starting audit report generation...")
            self.log(f"📊 Countsheet: {self.countsheet_path.get().split('/')[-1]}")
            self.log(f"📋 Template: {os.path.basename(self.template_path)}")  # Use resource_path
            self.log(f"📁 Master: {self.master_path.get().split('/')[-1]}")
            self.log("=" * 60)
            
            # Redirect stdout to capture print statements
            old_stdout = sys.stdout
            sys.stdout = mystdout = io.StringIO()
            
            try:
                # Call the main processing function
                self.process_audit_report()
                
                # Get captured output
                output = mystdout.getvalue()
                if output:
                    for line in output.split('\n'):
                        if line.strip():
                            self.log(f"📋 {line}")
                
                self.log("=" * 60)
                self.log("✅ Report generated successfully!")
                self.log(f"💾 Saved to: {self.output_path.get().split('/')[-1]}")
                
                # Show success message
                self.root.after(0, lambda: messagebox.showinfo("Success", 
                    "Audit report generated successfully!\n\n" +
                    f"Output saved to:\n{self.output_path.get()}"))
                
            finally:
                sys.stdout = old_stdout
                
        except Exception as e:
            self.log(f"❌ Error: {str(e)}")
            self.root.after(0, lambda e=e: messagebox.showerror("Error", 
                f"Failed to generate report:\n\n{str(e)}"))
        
        finally:
            # Reset UI state
            self.processing = False
            self.root.after(0, self.reset_ui_state)
    
    def reset_ui_state(self):
        """Reset UI to normal state after processing"""
        self.generate_btn.config(text="🚀 Generate Report", state='normal', bg='#89b4fa')
        self.progress.stop()
    
    def process_audit_report(self):
        """Main processing logic - adapted from your original script"""
        
        # File paths from GUI
        COUNTSHEET_FILE = self.countsheet_path.get()
        TEMPLATE_FILE = self.template_path  # Use resource_path
        MASTER_FILE = self.master_path.get()
        OUTPUT_FILE = self.output_path.get()
        
        # Your original mapping
        MAPPING = {
            0: lambda i, row: i+1,  # Sr No
            2: lambda i, row: 'Rutul Shah Co & LLp',  # Audit Team (static)
            3: lambda i, row: row.get('Distributor code', ''),  # Anchor Code
            5: lambda i, row: row.get('Distributor Name', ''),  # Distributor name
            6: lambda i, row: row.get('Item/SKU Code', ''),  # Article Code
            7: lambda i, row: row.get('Item Name', ''),  # Brand Pack
            8: lambda i, row: row.get('Field 2', ''),  # NPI / NON - NPI
            9: lambda i, row: row.get('Field 3', ''),  # Rate Excluding GST
            10: lambda i, row: row.get('Item Rate', ''),  # Rate Including GST
            11: lambda i, row: row.get('Field 1', ''),  # GST (%)
            12: lambda i, row: row.get('Field 4', ''),  # Standard Pack
            13: lambda i, row: row.get('Original QTY', ''),  # Primary Damage (Pcs)
            14: lambda i, row: row.get('Original Damage', ''),  # Non-Saleable product and Non-manufacturing Defect (Pcs)
            15: lambda i, row: row.get('Original Expired', ''),  # BBD Stock (Pcs)
            21: lambda i, row: row.get('Manu Date', ''),  # Manufacturing Date
            22: lambda i, row: row.get('Expiry Date', ''),  # Expiry Date
            25: lambda i, row: row.get('Remarks', ''),  # Remarks
        }
        
        # Read Countsheet data
        print("Loading countsheet data...")
        wb_count = openpyxl.load_workbook(COUNTSHEET_FILE, data_only=True)
        ws_count = wb_count.active
        count_headers = [cell.value for cell in next(ws_count.iter_rows(min_row=1, max_row=1))]
        count_data = []
        
        for row in ws_count.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(count_headers, row))
            
            def is_valid(val):
                if val is None:
                    return False
                if isinstance(val, str):
                    return val.strip() not in ("", "0", "0.0")
                try:
                    return float(val) != 0.0
                except Exception:
                    return True
            
            values = [row_dict.get(col) for col in ['Original QTY', 'Original Damage', 'Original Expired']]
            if any(is_valid(v) for v in values):
                count_data.append(row_dict)
        
        print(f"Found {len(count_data)} valid data rows")
        
        # Open template and get sheet
        print("Loading template...")
        wb_temp = openpyxl.load_workbook(TEMPLATE_FILE)
        ws_temp = wb_temp['Artical level format']
        
        # Store formatting and formula from original row 5
        formatting = []
        for col in range(1, 28):
            cell = ws_temp.cell(row=5, column=col)
            formula = None
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
            formatting.append({
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'number_format': cell.number_format,
                'protection': copy(cell.protection),
                'alignment': copy(cell.alignment),
                'formula': formula
            })
        
        # Find the row with the 'Total' label
        total_label_row = None
        for row in range(1, ws_temp.max_row + 1):
            cell = ws_temp.cell(row=row, column=5)
            if cell.value and isinstance(cell.value, str) and 'total' in cell.value.lower():
                total_label_row = row
                break
        
        if total_label_row is None:
            raise Exception("Could not find the 'Total' label row in column E.")
        
        # Remove existing data rows and insert new ones
        if total_label_row > 5:
            ws_temp.delete_rows(5, total_label_row - 5)
        
        ws_temp.insert_rows(5, len(count_data))
        
        print("Formatting and populating data rows...")
        
        # Apply formatting and populate data
        for i in range(len(count_data)):
            for col in range(1, 28):
                cell = ws_temp.cell(row=5 + i, column=col)
                fmt = formatting[col-1]
                cell.font = copy(fmt['font'])
                cell.border = copy(fmt['border'])
                cell.fill = copy(fmt['fill'])
                cell.number_format = fmt['number_format']
                cell.protection = copy(fmt['protection'])
                cell.alignment = copy(fmt['alignment'])
                cell.value = None
        
        # Write data and formulas
        for i, row in enumerate(count_data):
            for col_idx in range(27):
                cell = ws_temp.cell(row=5 + i, column=col_idx+1)
                if col_idx in MAPPING:
                    value = MAPPING[col_idx](i, row)
                    cell.value = value
                else:
                    fmt = formatting[col_idx]
                    if fmt['formula']:
                        def repl(m):
                            col_letter = m.group(1)
                            return f"{col_letter}{5 + i}"
                        formula = re.sub(r'([A-Z]+)5', repl, fmt['formula'])
                        cell.value = formula
            
            # Calculate INR columns
            try:
                rate_incl_gst = float(row.get('Item Rate', 0) or 0)
                primary_damage_pcs = float(row.get('Original QTY', 0) or 0)
                non_saleable_pcs = float(row.get('Original Damage', 0) or 0)
                bbd_stock_pcs = float(row.get('Original Expired', 0) or 0)
            except:
                rate_incl_gst = primary_damage_pcs = non_saleable_pcs = bbd_stock_pcs = 0
            
            # Calculate totals
            total_verified_qty = primary_damage_pcs + non_saleable_pcs + bbd_stock_pcs
            ws_temp.cell(row=5 + i, column=17).value = total_verified_qty
            
            primary_damage_inr = primary_damage_pcs * rate_incl_gst
            non_saleable_inr = non_saleable_pcs * rate_incl_gst
            bbd_stock_inr = bbd_stock_pcs * rate_incl_gst
            
            ws_temp.cell(row=5 + i, column=18).value = primary_damage_inr
            ws_temp.cell(row=5 + i, column=19).value = non_saleable_inr
            ws_temp.cell(row=5 + i, column=20).value = bbd_stock_inr
            
            total_audited_value = primary_damage_inr + non_saleable_inr + bbd_stock_inr
            ws_temp.cell(row=5 + i, column=21).value = total_audited_value
        
        print("Calculating subtotals...")
        
        # Calculate subtotals
        subtotal_row = 5 + len(count_data)
        
        def safe_float(val):
            try:
                return float(val)
            except (TypeError, ValueError):
                return 0.0
        
        for col in range(14, 22):
            col_values = []
            for i in range(len(count_data)):
                val = ws_temp.cell(row=5 + i, column=col).value
                if isinstance(val, str) and val.startswith('='):
                    continue
                col_values.append(safe_float(val))
            subtotal = sum(col_values)
            ws_temp.cell(row=subtotal_row, column=col).value = subtotal
        
        # Save initial file
        wb_temp.save(OUTPUT_FILE)
        print(f'Data written to {OUTPUT_FILE}')
        
        # Master file lookup
        print("Processing master file lookup...")
        
        wb_out = openpyxl.load_workbook(OUTPUT_FILE)
        ws_out = wb_out['Artical level format']
        anchor_code = ws_out.cell(row=5, column=4).value
        distributor_name = ws_out.cell(row=5, column=6).value
        
        print(f"Looking up: Anchor Code = {anchor_code}, Distributor = {distributor_name}")
        print(f"DEBUG: Anchor Code type: {type(anchor_code)}, value: '{anchor_code}'")
        print(f"DEBUG: Distributor Name type: {type(distributor_name)}, value: '{distributor_name}'")
        
        wb_master = openpyxl.load_workbook(MASTER_FILE, data_only=True)
        ws_master = wb_master.active
        master_headers = [cell.value for cell in next(ws_master.iter_rows(min_row=1, max_row=1))]
        print(f"DEBUG: Master file headers: {master_headers}")
        
        # Find relevant columns
        ac_idx = None
        dn_idx = None
        for idx, h in enumerate(master_headers):
            if h and 'anchor code' in str(h).lower():
                ac_idx = idx
                print(f"DEBUG: Found Anchor Code column at index {idx}: '{h}'")
            if h and ('db name' in str(h).lower() or 'distributor name' in str(h).lower()):
                dn_idx = idx
                print(f"DEBUG: Found Distributor Name column at index {idx}: '{h}'")
        
        if ac_idx is None or dn_idx is None:
            print('Could not find required columns in master file')
            print(f"DEBUG: ac_idx = {ac_idx}, dn_idx = {dn_idx}")
            return
        
        # Add normalize_code function for robust matching
        def normalize_code(val):
            s = str(val).strip()
            if s.endswith('.0'):
                s = s[:-2]
            return s
        
        # Find matching row
        found = False
        print(f"DEBUG: Searching for Anchor Code '{anchor_code}' and Distributor '{distributor_name}'")
        row_count = 0
        for row in ws_master.iter_rows(min_row=2, values_only=True):
            row_count += 1
            if row_count <= 5:  # Show first 5 rows for debugging
                print(f"DEBUG: Row {row_count}: AC='{row[ac_idx]}', DN='{row[dn_idx]}'")
            if normalize_code(row[ac_idx]) == normalize_code(anchor_code) and \
               str(row[dn_idx]).strip().lower() == str(distributor_name).strip().lower():
                row_dict = dict(zip(master_headers, row))
                print('Found matching row in master file')
                print(f"DEBUG: Match found at row {row_count}")
                anchor_name_val = row_dict.get('Anchor Name')
                region_val = row_dict.get('Region')
                
                # Find audit serial number
                audit_key = next((h for h in master_headers 
                                if h and str(h).strip().lower() == "audit std serial no"), None)
                audit_std_serial_no_val = row[master_headers.index(audit_key)] if audit_key else ""
                
                # Fill data in output file
                for i in range(len(count_data)):
                    ws_out.cell(row=5 + i, column=5).value = anchor_name_val
                    ws_out.cell(row=5 + i, column=2).value = region_val
                
                ws_out.cell(row=2, column=2).value = audit_std_serial_no_val
                
                # Fill Sign Format sheet
                print("Updating Sign Format sheet...")
                ws_sign = wb_out['Sign Format.']
                
                ws_sign.cell(row=4, column=3).value = audit_std_serial_no_val
                ws_sign.cell(row=6, column=3).value = row_dict.get('Anchor Code/ DB Code', '')
                
                db_name = row_dict.get('Anchor Name', '')
                ws_sign.cell(row=7, column=3).value = db_name

                dist_name = row_dict.get('DB Name', '')
                #ws_sign.cell(row=7, column=3).value = dist_name
                
                city = row_dict.get('Distributor City', '')
                ws_sign.cell(row=8, column=3).value = f"{dist_name} & {city}" if dist_name or city else ''
                ws_sign.cell(row=9, column=3).value = datetime.datetime.now().strftime('%d-%m-%Y')
                
                # Update date placeholders
                today_str = datetime.datetime.now().strftime('%d-%m-%Y')
                
                for cell_ref in ['B20', 'B22']:
                    note_cell = ws_sign[cell_ref]
                    if note_cell.value:
                        if 'date upto  ( )' in note_cell.value:
                            note_cell.value = note_cell.value.replace('date upto  ( )', f'date upto ({today_str})')
                        elif 'date of Audit. ( )' in note_cell.value:
                            note_cell.value = note_cell.value.replace('date of Audit. ( )', f'date of Audit. ({today_str})')
                        elif '()' in note_cell.value:
                            note_cell.value = note_cell.value.replace('()', f'({datetime.datetime.now().strftime("%d-%m-%y")})')
                
                # Add subtotals to Sign Format
                ws_art = wb_out['Artical level format']
                total_row_idx = None
                for row_idx in range(1, ws_art.max_row + 1):
                    cell = ws_art.cell(row=row_idx, column=5)
                    if cell.value and isinstance(cell.value, str) and 'total' in cell.value.lower():
                        total_row_idx = row_idx
                        break
                
                if total_row_idx:
                    ws_sign.cell(row=12, column=2).value = ws_art.cell(row=total_row_idx, column=14).value
                    ws_sign.cell(row=12, column=3).value = ws_art.cell(row=total_row_idx, column=15).value
                    ws_sign.cell(row=12, column=4).value = ws_art.cell(row=total_row_idx, column=16).value
                    ws_sign.cell(row=12, column=6).value = ws_art.cell(row=total_row_idx, column=18).value
                    ws_sign.cell(row=12, column=7).value = ws_art.cell(row=total_row_idx, column=19).value
                    ws_sign.cell(row=12, column=8).value = ws_art.cell(row=total_row_idx, column=20).value
                    ws_sign.cell(row=12, column=10).value = row_dict.get('Reported Value', '')
                
                found = True
                break
        
        if not found:
            print('No matching row found in master file')
        
        # --- Group by Manufacturing Quarter and aggregate Total Audited Value ---
        print("Processing Manufacturing Quarter grouping...")
        
        # Read data from Artical level format sheet with calculated values (not formulas)
        wb_art_data = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
        ws_art = wb_art_data['Artical level format']
        
        # Manufacturing Date is in column V (22) and Total Audited Value is in column U (21)
        manu_date_col = 22  # Column V (Manufacturing Date)
        total_audited_value_col = 21  # Column U (Total Audited Value Including GST)
        
        print(f"Using Manufacturing Date column V ({manu_date_col})")
        print(f"Using Total Audited Value column U ({total_audited_value_col})")
        
        def calculate_quarter(manufacturing_date):
            """Calculate quarter based on manufacturing date"""
            if not manufacturing_date:
                return None
            
            try:
                # Convert to datetime if it's a string
                if isinstance(manufacturing_date, str):
                    # Try different date formats
                    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                        try:
                            manufacturing_date = datetime.datetime.strptime(manufacturing_date, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        return None
                
                # If it's already a datetime object
                if isinstance(manufacturing_date, datetime.datetime):
                    month = manufacturing_date.month
                    year = manufacturing_date.year
                    
                    # Calculate fiscal year (April to March) - use short format (24-25)
                    if month >= 4:  # April to December
                        fiscal_year = f"{str(year)[-2:]}-{str(year+1)[-2:]}"
                    else:  # January to March
                        fiscal_year = f"{str(year-1)[-2:]}-{str(year)[-2:]}"
                    
                    # Calculate quarter
                    if month in [4, 5, 6]:
                        quarter = "Q1"
                    elif month in [7, 8, 9]:
                        quarter = "Q2"
                    elif month in [10, 11, 12]:
                        quarter = "Q3"
                    else:  # 1, 2, 3
                        quarter = "Q4"
                    
                    return f"{quarter} FY {fiscal_year}"
                
                return None
            except Exception as e:
                print(f"Error calculating quarter for date {manufacturing_date}: {e}")
                return None
        
        # Group by Manufacturing Quarter and aggregate Total Audited Value
        quarter_data = {}
        
        # Read data rows (starting from row 5)
        for row in range(5, ws_art.max_row + 1):
            manu_date_cell = ws_art.cell(row=row, column=manu_date_col)
            value_cell = ws_art.cell(row=row, column=total_audited_value_col)
            
            manu_date = manu_date_cell.value
            value = value_cell.value
            
            # Calculate quarter from manufacturing date
            quarter = calculate_quarter(manu_date)
            
            # Debug: Print first few rows to see what we're reading
            if row <= 10:
                print(f"Row {row}: Date='{manu_date}' (type: {type(manu_date)}), Quarter='{quarter}', Value='{value}' (type: {type(value)})")
            
            # Skip if quarter is empty or None
            if not quarter or quarter == '':
                continue
            
            # Convert value to float, skip if not numeric
            try:
                value = float(value) if value is not None else 0.0
            except (ValueError, TypeError):
                value = 0.0
            
            # Aggregate by quarter
            if quarter in quarter_data:
                quarter_data[quarter] += value
            else:
                quarter_data[quarter] = value
        
        print(f"Found {len(quarter_data)} unique quarters: {list(quarter_data.keys())}")
        
        # Sort quarters (assuming they are in format like "Q1 FY 2024-2025", "Q2 FY 2024-2025", etc.)
        sorted_quarters = sorted(quarter_data.keys())
        
        # Clear C16-I16 and C17-I17 first
        for col in range(3, 10):  # C to I (columns 3-9)
            ws_sign.cell(row=16, column=col).value = None
            ws_sign.cell(row=17, column=col).value = None
        
        # Fill quarter names in C16-I16 and aggregated values in C17-I17
        for i, quarter in enumerate(sorted_quarters):
            if i >= 7:  # Maximum 7 columns (C to I)
                break
            
            col = 3 + i  # Start from column C (3)
            aggregated_value = quarter_data[quarter]
            
            # Fill quarter name in row 16
            ws_sign.cell(row=16, column=col).value = quarter
            # Fill aggregated value in row 17
            ws_sign.cell(row=17, column=col).value = aggregated_value
            
            print(f"Filled {quarter}: {aggregated_value} in column {get_column_letter(col)}")
        
        print(f"Manufacturing Quarter grouping completed. Filled {len(sorted_quarters)} quarters.")
        
        # Save final file
        wb_out.save(OUTPUT_FILE)
        print(f'Report completed and saved to {OUTPUT_FILE}')


def main():
    root = tk.Tk()
    app = AuditReportGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
